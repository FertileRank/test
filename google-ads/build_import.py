#!/usr/bin/env python3
"""
Build the EggCelle Google Ads Editor Import (June 2026 optimization).

Generates Google Ads Editor-compatible CSV files plus a consolidated, review-
friendly XLSX workbook that encodes the audit recommendations focused on
LOWERING COST PER LEAD (Submit Lead Form) and reallocating budget toward the
2x-more-efficient SKAG campaigns.

Data sources:
  - EggCelle Senior Google Ads Audit (June 2026)
  - Google Ads Editor export (7 campaigns / 374 ad groups, 2026-05-31)
  - Live Search Atlas PPC data, account 12070 (cid 6026920997), 2026-05-10..06-09
  - Search Atlas keyword CPC/volume benchmarks
"""
import csv, os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE = os.path.dirname(os.path.abspath(__file__))
IMP = os.path.join(BASE, "import")
os.makedirs(IMP, exist_ok=True)

# Exact campaign names as they exist in Google Ads Editor (from the export file).
MI   = "FertiRank - Ads - EggCelle — Michigan"
NC   = "FertiRank - Ads - EggCelle — North Carolina"
OH   = "FertiRank - Ads - EggCelle — Ohio"
OTTO = "OTTO - Ads - RMA Max Conversion"
MISK = "FertiRank - Ads - EggCelle | NB | MI | SKAG"
OHSK = "FertiRank - Ads - EggCelle | NB | OH | SKAG"
NCSK = "FertiRank - Ads - EggCelle | NB | NC | SKAG"

# Current state (confirmed by audit + live API, May 10-Jun 9 2026).
# name: (current_budget, current_real_CPA, current_conv, current_bid_strategy)
CURRENT = {
    MI:   (350, 38.43, 202, "Maximize conversions"),
    NC:   (350, 43.31, 216, "Maximize conversions"),
    OH:   (250, 30.95, 179, "Maximize conversions"),
    OTTO: (50,  18.33, 24,  "Maximize conversions"),
    MISK: (50,  14.31, 57,  "Maximize clicks"),
    OHSK: (25,  21.09, 2,   "Maximize clicks"),
    NCSK: (50,  20.29, 140, "Maximize clicks"),
}

# Proposed campaign-level changes.
# name: dict(new_budget, new_tcpa(None=leave), set_lead_goal(bool), note)
PROPOSED = {
    NC:   dict(budget=275, tcpa=37, lead_goal=True,
               note="Highest-CPA market ($43). Trim & shift to NC SKAG (same geo, ~$20 CPA). Tighten tCPA."),
    MI:   dict(budget=300, tcpa=35, lead_goal=True,
               note="Trim modestly; add tCPA to control cost. Fund MI SKAG (best efficiency)."),
    OH:   dict(budget=240, tcpa=28, lead_goal=True,
               note="Most efficient market ($31) & rank-limited — trim least, add tCPA."),
    OTTO: dict(budget=35,  tcpa=18, lead_goal=True,
               note="Pending conversion-tracking audit. Switch goal to Submit Lead Form; hold tCPA ~$18."),
    MISK: dict(budget=100, tcpa=None, lead_goal=False,
               note="Best efficiency ($14 CPA). PRESERVE bid strategy (audit D2/winner). Budget +$50 to scale."),
    OHSK: dict(budget=75,  tcpa=None, lead_goal=False,
               note="Underfunded (2 conv). Raise to $75 to exit learning & gather data (audit O2)."),
    NCSK: dict(budget=150, tcpa=None, lead_goal=False,
               note="Crown jewel ($20 CPA, losing 7.6% IS to budget). PRESERVE settings; budget $50->$150 (audit O1)."),
}

ORDER = [NC, MI, OH, OTTO, NCSK, MISK, OHSK]

# ----- Cross-market geo negatives (campaign-level, Negative Phrase) -----
GEO = {
    "MI": ["michigan", "detroit", "troy", "bloomfield hills", "royal oak"],
    "NC": ["north carolina", "charlotte", "raleigh", "greensboro", "winston-salem"],
    "OH": ["ohio", "columbus", "cincinnati", "dayton", "cleveland"],
}
# For each state's campaigns, exclude the OTHER two states' geo terms.
STATE_CAMPAIGNS = {"MI": [MI, MISK], "NC": [NC, NCSK], "OH": [OH, OHSK]}
OTHER = {"MI": ["NC", "OH"], "NC": ["MI", "OH"], "OH": ["MI", "NC"]}

neg_rows = []
for st, camps in STATE_CAMPAIGNS.items():
    block = []
    for o in OTHER[st]:
        block += GEO[o]
    for camp in camps:
        for term in block:
            neg_rows.append({
                "Campaign": camp, "Ad Group": "",
                "Keyword": term, "Criterion Type": "Negative Phrase",
                "Campaign Status": "Enabled",
            })

# =========================================================================
# 1) CSV: campaign budget changes (all 7) -- zero-risk, applies independently
# =========================================================================
with open(os.path.join(IMP, "01_campaign_budget_changes.csv"), "w", newline="") as f:
    w = csv.writer(f)
    w.writerow(["Campaign", "Budget", "Budget type", "Campaign Status"])
    for c in ORDER:
        w.writerow([c, f"{PROPOSED[c]['budget']:.2f}", "Daily", "Enabled"])

# =========================================================================
# 2) CSV: Market + OTTO -> Target CPA + Submit Lead Form conversion goal
# =========================================================================
with open(os.path.join(IMP, "02_market_targetcpa_and_conversion_goals.csv"), "w", newline="") as f:
    w = csv.writer(f)
    w.writerow(["Campaign", "Bid Strategy Type", "Target CPA",
                "Standard conversion goals", "Custom conversion goal", "Campaign Status"])
    for c in [NC, MI, OH, OTTO]:
        p = PROPOSED[c]
        w.writerow([c, "Maximize conversions", f"{p['tcpa']:.2f}",
                    "Submit lead form(Website)", "", "Enabled"])

# =========================================================================
# 3) CSV: cross-market negative keywords
# =========================================================================
with open(os.path.join(IMP, "03_cross_market_negative_keywords.csv"), "w", newline="") as f:
    w = csv.DictWriter(f, fieldnames=["Campaign", "Ad Group", "Keyword",
                                      "Criterion Type", "Campaign Status"])
    w.writeheader()
    for r in neg_rows:
        w.writerow(r)

print(f"CSVs written to {IMP}")
print(f"  01 budgets: {len(ORDER)} campaigns")
print(f"  02 tCPA+goals: 4 campaigns")
print(f"  03 negatives: {len(neg_rows)} rows")

# =========================================================================
# 4) Consolidated review workbook
# =========================================================================
wb = openpyxl.Workbook()

H_FILL = PatternFill("solid", fgColor="1F4E78")
H_FONT = Font(bold=True, color="FFFFFF", size=11)
SUB_FILL = PatternFill("solid", fgColor="D9E1F2")
GOOD = PatternFill("solid", fgColor="C6EFCE")
WARN = PatternFill("solid", fgColor="FFEB9C")
BAD  = PatternFill("solid", fgColor="FFC7CE")
TITLE = Font(bold=True, size=14, color="1F4E78")
BOLD = Font(bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_header(ws, row=1, ncols=10):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = H_FILL; cell.font = H_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = BORDER

def setw(ws, widths):
    for col, wid in widths.items():
        ws.column_dimensions[col].width = wid

# ---- Sheet: READ ME ----
ws = wb.active
ws.title = "READ ME"
setw(ws, {"A": 3, "B": 100})
rows = [
    ("EggCelle — Google Ads Editor Import (June 2026 Optimization)", TITLE),
    ("Goal: lower cost per LEAD (Submit Lead Form) and scale the 2x-efficient SKAG campaigns.", BOLD),
    ("", None),
    ("Prepared from: June 2026 Senior Audit + live Search Atlas account data (acct 12070 / cid 6026920997,", None),
    ("window 2026-05-10 to 2026-06-09) + Search Atlas keyword CPC benchmarks.", None),
    ("", None),
    ("HOW TO IMPORT", BOLD),
    ("1. Google Ads Editor → Account → Import → 'From file…'  (back up / 'Get latest changes' first).", None),
    ("2. Import the CSVs in /import in order. Each is independent — review the preview before 'Keep'.", None),
    ("     • 01_campaign_budget_changes.csv  — new daily budgets (all 7 campaigns)", None),
    ("     • 02_market_targetcpa_and_conversion_goals.csv — Target CPA + Submit-Lead-Form goal (Market+OTTO)", None),
    ("     • 03_cross_market_negative_keywords.csv — geo negatives (60) to stop cross-state spend leakage", None),
    ("3. Review every change in Editor's preview, then Post to the account.", None),
    ("", None),
    ("WHY THIS LOWERS COST", BOLD),
    ("• The account counts 'Outbound Click' as a conversion goal alongside 'Submit Lead Form'. Market & OTTO", None),
    ("  campaigns optimize to a custom 'Conversion Funnel' goal that includes those clicks — inflating", None),
    ("  conversions and pulling bids toward cheap clicks instead of real leads. Switching them to the", None),
    ("  'Submit Lead Form' goal refocuses bidding on actual donor leads. (Audit D1/T1.)", None),
    ("• SKAG campaigns already optimize to Submit Lead Form and run ~$14–20 CPA vs $31–43 for Market.", None),
    ("  We shift ~$200/day of marginal budget from high-CPA Market inventory into the SKAGs (net budget", None),
    ("  change only +$50/day) — the single biggest blended-CPA lever. (Audit O1/O2.)", None),
    ("• Target CPA caps are added to Market + OTTO ~8–15% below current CPA to tighten cost.", None),
    ("• 60 cross-market geo negatives stop MI/NC/OH campaigns paying for each other's city/state queries. (O5)", None),
    ("", None),
    ("WHAT IS NOT IN THIS IMPORT (do in the Google Ads UI — see ChangeLog .md)", BOLD),
    ("• Conversion-tracking audit: set 'Outbound Click' to Secondary (not a conversion goal) at account level.", None),
    ("• Offline conversion import (prescreen→approval) — highest-leverage fix; aligns spend to qualified donors.", None),
    ("• Reactivate Brand Protection (a paused campaign not in this 7-campaign export).", None),
    ("• Phase 2 (after this stabilizes): move SKAGs Max-Clicks→Max-Conversions; prune zero-conversion ad groups.", None),
    ("", None),
    ("NOTE: NC & MI SKAG bid strategies are intentionally NOT changed here — the audit flags them as winners", None),
    ("to preserve (D2). We only raise their budgets to avoid disrupting their learning.", None),
]
r = 1
for text, font in rows:
    cell = ws.cell(row=r, column=2, value=text)
    if font: cell.font = font
    r += 1

# ---- Sheet: Current vs Proposed ----
ws = wb.create_sheet("Current vs Proposed")
hdr = ["Campaign", "Bucket", "Current Budget/day", "New Budget/day", "Δ Budget",
       "Bid Strategy", "New Target CPA", "Current CPA (real lead)", "Conversion Goal change", "Rationale"]
ws.append(hdr); style_header(ws, 1, len(hdr))
bucket = {NC:"NB Market", MI:"NB Market", OH:"NB Market", OTTO:"Legacy",
          NCSK:"NB SKAG", MISK:"NB SKAG", OHSK:"NB SKAG"}
for c in ORDER:
    cur = CURRENT[c]; p = PROPOSED[c]
    delta = p["budget"] - cur[0]
    goal = "→ Submit Lead Form only" if p["lead_goal"] else "(already Submit Lead Form)"
    bidtxt = cur[3] + (" + tCPA" if p["tcpa"] else " (unchanged)")
    ws.append([c, bucket[c], cur[0], p["budget"], f"{delta:+d}", bidtxt,
               (f"${p['tcpa']}" if p["tcpa"] else "—"),
               f"${cur[1]:.2f}", goal, p["note"]])
# totals
cb = sum(CURRENT[c][0] for c in ORDER); nb = sum(PROPOSED[c]["budget"] for c in ORDER)
ws.append(["TOTAL (active daily)", "", cb, nb, f"{nb-cb:+d}", "", "", "", "",
           f"Monthly: ${cb*30:,} → ${nb*30:,} (≈{(nb-cb)*30:+,}/mo)"])
last = ws.max_row
for c in range(1, len(hdr)+1):
    ws.cell(row=last, column=c).font = BOLD
    ws.cell(row=last, column=c).fill = SUB_FILL
for rr in range(2, ws.max_row+1):
    for c in range(1, len(hdr)+1):
        ws.cell(row=rr, column=c).border = BORDER
        ws.cell(row=rr, column=c).alignment = WRAP
    # color delta + bucket
    dcell = ws.cell(row=rr, column=5)
    try:
        dv = int(str(dcell.value).replace("+",""))
        dcell.fill = GOOD if dv > 0 else (BAD if dv < 0 else WARN)
    except (ValueError, TypeError):
        pass
setw(ws, {"A":42,"B":11,"C":15,"D":14,"E":10,"F":24,"G":13,"H":18,"I":24,"J":58})
ws.freeze_panes = "A2"

# ---- Sheet: Campaign (import-ready, budgets) ----
ws = wb.create_sheet("Campaign budgets (import)")
h = ["Campaign", "Budget", "Budget type", "Campaign Status"]
ws.append(h); style_header(ws, 1, len(h))
for c in ORDER:
    ws.append([c, f"{PROPOSED[c]['budget']:.2f}", "Daily", "Enabled"])
setw(ws, {"A":46,"B":12,"C":12,"D":16})
ws.freeze_panes = "A2"

# ---- Sheet: Market tCPA + goals (import-ready) ----
ws = wb.create_sheet("Market tCPA + goals (import)")
h = ["Campaign", "Bid Strategy Type", "Target CPA", "Standard conversion goals",
     "Custom conversion goal", "Campaign Status"]
ws.append(h); style_header(ws, 1, len(h))
for c in [NC, MI, OH, OTTO]:
    p = PROPOSED[c]
    ws.append([c, "Maximize conversions", f"{p['tcpa']:.2f}",
               "Submit lead form(Website)", "", "Enabled"])
setw(ws, {"A":46,"B":22,"C":12,"D":26,"E":22,"F":16})
ws.freeze_panes = "A2"

# ---- Sheet: Negative keywords (import-ready) ----
ws = wb.create_sheet("Negative keywords (import)")
h = ["Campaign", "Ad Group", "Keyword", "Criterion Type", "Campaign Status"]
ws.append(h); style_header(ws, 1, len(h))
for rrow in neg_rows:
    ws.append([rrow["Campaign"], rrow["Ad Group"], rrow["Keyword"],
               rrow["Criterion Type"], rrow["Campaign Status"]])
setw(ws, {"A":46,"B":10,"C":20,"D":16,"E":16})
ws.freeze_panes = "A2"

# ---- Sheet: Keyword benchmarks (Search Atlas) ----
ws = wb.create_sheet("Keyword benchmarks (SA)")
h = ["Keyword theme", "Monthly Volume (US)", "Avg CPC", "Intent", "In which campaigns", "Read"]
ws.append(h); style_header(ws, 1, len(h))
kb = [
    ("egg donor requirements", 5400, 3.97, "commercial", "SKAG REQS / Market", "High vol, LOW CPC — efficient. Lean in."),
    ("requirements to become an egg donor", 5400, 3.91, "commercial", "SKAG REQS", "Efficient long-tail."),
    ("how much do egg donors get paid", 1650, 3.75, "informational", "SKAG COMP", "Low CPC; strong SKAG converter."),
    ("donate eggs for money", 3600, 5.61, "commercial", "SKAG MONEY", "Core money intent."),
    ("get paid to donate eggs", 220, 7.16, "commercial", "SKAG MONEY", "Exact converter in MI/NC SKAG."),
    ("egg donation near me", 12100, 8.68, "commercial", "Market LOCAL", "HIGH CPC head term — drives Market CPA up."),
    ("egg donation", 22200, 8.52, "commercial", "Market", "HIGH CPC, broad — keep tight via tCPA/negatives."),
    ("become an egg donor", 798, 12.33, "informational", "Market/SKAG BECOME", "Most expensive CPC — cap with tCPA."),
]
for row in kb:
    ws.append(list(row))
for rr in range(2, ws.max_row+1):
    cpc = ws.cell(row=rr, column=3).value
    fill = GOOD if cpc < 5 else (WARN if cpc < 8 else BAD)
    ws.cell(row=rr, column=3).fill = fill
    for c in range(1, len(h)+1):
        ws.cell(row=rr, column=c).border = BORDER
        ws.cell(row=rr, column=c).alignment = WRAP
ws.append([])
ws.append(["Insight: efficient SKAGs win on low-CPC 'requirements / compensation / money' long-tail ($3.75–5.61).",
           "", "", "", "", ""])
ws.cell(row=ws.max_row, column=1).font = BOLD
ws.append(["Costly Market campaigns lean on high-CPC head terms ($8.52–12.33). Shifting budget to SKAGs lowers blended CPA.",
           "", "", "", "", ""])
ws.cell(row=ws.max_row, column=1).font = BOLD
setw(ws, {"A":40,"B":18,"C":10,"D":15,"E":22,"F":52})
ws.freeze_panes = "A2"

OUT = os.path.join(BASE, "EggCelle_GoogleAds_Editor_Import_June2026.xlsx")
wb.save(OUT)
print(f"Workbook written: {OUT}")
